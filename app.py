import pandas as pd
import openpyxl
import PyPDF2
from flask import Flask, request, jsonify
from flask_cors import CORS
from llama_index import Document, GPTTreeIndex
import os
import json
import base64
import io

with open('config.json', 'r') as f:
    config = json.load(f)

os.environ['OPENAI_API_KEY'] = config['api_key']

app = Flask(__name__)
CORS(app)
index = None

@app.route('/initialize', methods=['POST'])
def initialize_index():
    global index

    context = ''
    
    uploaded_files = request.files.getlist("files")
    for file in uploaded_files:
        file.save(file.filename)
        print (file.filename)

        def getTextFromExcel(filename):
            text = ''

            print (filename)
            wb = openpyxl.load_workbook(filename)
            for ws in wb.worksheets:
                for row in range(1, ws.max_row+1):
                    solution_title = ws['A' + str(row)].value
                    solution_details = ws['B' + str(row)].value
                    text = text + str(solution_title) + str(solution_details)
            return text
        context = context + getTextFromExcel(file.filename)
    
    document = [Document(context)]
    index = GPTTreeIndex(document)
    return "done"

@app.route('/query', methods=['POST'])
def query_index():
    query_data = request.get_json()
    query = query_data['query']
    response = index.query(query)
    return jsonify({"response" : response.response})

if __name__ == '__main__':
    app.run(host="127.0.0.1")


        # file = request.files['file']
    # print(request.files)
    # print(file)
    # file.save(file.filename)
    # return "done"

    # data = request.get_json()
    # base64_string = data['pdf']
    # decode = base64.b64decode(base64_string)
    # excel_file = io.BytesIO(decode)

    # pdf_reader = PyPDF2.PdfReader(excel_file)
    # num_pages = len(pdf_reader.pages)
    # pdf_text = ''

    # for page in range(num_pages):
    #     page_obj = pdf_reader.pages[page]
    #     text = page_obj.extract_text()
    #     pdf_text += text
    # excel_file.close()

    # ps = openpyxl.load_workbook(excel_file)



    # ps = openpyxl.load_workbook('Solutions.xlsx')
    # sheet = ps['Sheet1']
    # text = ''
    # for row in range(1, sheet.max_column+1):
    #     solution_title = sheet['A' + str(row)].value
    #     solution_details = sheet['B' + str(row)].value
    #     text += solution_title
    #     text += solution_details



    # document = [Document(text)]
    # index = GPTTreeIndex(document)
    # return jsonify({"response": "This is response from"})